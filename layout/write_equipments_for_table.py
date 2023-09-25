from pandas import isna
import numpy

from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string

from fastnumbers import fast_real, try_int, try_forceint, try_real

from data_frame_features import EquipmentsSourceData
from file_features.message import output_message
from data_frame_features import filter_data_frame
from layout.set_cell_style import set_cell_style
from layout.layout_setting import basic_colors, headers


def _equipment_line_output(equipment_info: numpy.record, sheet: worksheet, row: int,
                           attributes: dict[str:str] = None, options: dict[str:str] = None,
                           table_attributes: list[str] = None, table_options: list[str] = None) -> int:
    if equipment_info:
        columns = headers['table']
        for i in range(len(columns) - 1):
            value = equipment_info[columns[i + 1]]
            value = "" if isna(value) else value
            sheet.cell(row=row, column=column_index_from_string(columns[i])).value = value
            set_cell_style(sheet, row, column_index_from_string(columns[i]), 'equipment')

        stat = try_forceint(equipment_info['G'])
        if isinstance(stat, (int, float)) and stat == 0:
            stat = ''
        sheet.cell(row=row, column=column_index_from_string('F')).value = stat

        # раскрашиваем
        sheet.cell(row=row, column=column_index_from_string('A')).font = basic_colors['grey']
        value = equipment_info['C']
        value = "" if isna(value) else value.strip()
        color = basic_colors['black'] if value == "ТСН" else basic_colors['dark_red']
        sheet.cell(row=row, column=column_index_from_string('B')).font = color
        sheet.cell(row=row, column=column_index_from_string('C')).font = basic_colors['dark_blue_bold']
        sheet.cell(row=row, column=column_index_from_string('E')).font = basic_colors['dark_red']

        # атрибуты
        start_column = column_index_from_string('I')
        if attributes and len(attributes) > 0:
            for i, header_attribute in enumerate(table_attributes):
                value_out = attributes.get(header_attribute, " ")
                sheet.cell(row=row, column=start_column + i).value = value_out
                set_cell_style(sheet, row, start_column + i, style_name='equipment')

        # параметры
        start_column += len(table_attributes)
        # ставим стиль на всю строку
        if table_options and len(table_options) > 0:
            for column in range(start_column, start_column + 5 * len(table_options)):
                set_cell_style(sheet, row, column, style_name='equipment')

        if options and len(options) > 0:
            step = 0
            # перебираем параметры в шапке таблицы
            for i, header_option in enumerate(table_options):
                option_value = options.get(header_option, None)
                if option_value:
                    option_value = ["" if isna(value) else value for value in option_value]
                    column = step + i
                    sheet.cell(row=row, column=start_column + column).value = try_real(option_value[0])
                    sheet.cell(row=row, column=start_column + column + 1).value = try_real(option_value[1])
                    sheet.cell(row=row, column=start_column + column + 2).value = option_value[2]
                    sheet.cell(row=row, column=start_column + column + 3).value = try_real(option_value[3])
                    sheet.cell(row=row, column=start_column + column + 4).value = try_forceint(option_value[4])
                step += 4

        sheet.row_dimensions[row].height = 12
        sheet.row_dimensions[row + 1].height = 12
        return row + 1
    return row


def write_equipments_for_table(src_data: EquipmentsSourceData, sheet: worksheet, start_line: int,
                               table: numpy.record) -> int:
    """
    Записывает информацию о ресурсах для указанной таблицы.
    :param src_data: Данные.
    :param sheet: Лист, на который надо выводить данные.
    :param start_line: Строка с которой надо начинить запись.
    :param table: Данные о таблице.
    """
    table_code = table['C']

    # список атрибутов из шапки таблицы
    table_attributes = [x.strip() for x in table['G'].split(',')] if not isna(table['G']) else []
    # список параметров из шапки таблицы
    table_options = [x.strip() for x in table['H'].split(',')] if not isna(table['H']) else []

    row = start_line
    equipments_df = filter_data_frame(src_df=src_data.equipments, column_name="J", target=table_code)

    if not equipments_df.empty:
        for equipment in equipments_df.to_records(index=False):
            equipment_code = equipment["D"]

            attributes_df = filter_data_frame(src_df=src_data.attributes, column_name="B", target=equipment_code)
            attributes = {x[2]: x[3] for x in attributes_df.to_records(index=False).tolist()}

            options_df = filter_data_frame(src_df=src_data.options, column_name="B", target=equipment_code)
            options = {x[2]: x[3:] for x in options_df.to_records(index=False).tolist()}

            row = _equipment_line_output(equipment, sheet, row, attributes, options, table_attributes=table_attributes,
                                         table_options=table_options)
            del attributes_df
            del options_df
    else:
        output_message(f"у таблицы {table}", f"нет оборудования с Атрибутами/Параметрами.")

    del equipments_df
    return row
