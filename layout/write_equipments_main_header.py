from openpyxl.worksheet import worksheet

from layout.layout_setting import headers, width_columns, items_styles, basic_colors


def write_equipments_main_header(sheet: worksheet):
    sheet.append(["Оборудование"])
    sheet.append(headers["equipments_main"])
    for c in range(1, 4):
        sheet.cell(row=1, column=c).font = basic_colors['dark_red']
        sheet.cell(row=1, column=c).fill = items_styles['main_equipments_header']['fill']
        sheet.cell(row=1, column=c).border = items_styles['main_equipments_header']['border']

    for column in range(1, len(headers["equipments_main"]) + 1):
        sheet.cell(row=2, column=column).font = items_styles['main_equipments_header']['font']
        sheet.cell(row=2, column=column).fill = items_styles['main_equipments_header']['fill']
        sheet.cell(row=2, column=column).border = items_styles['main_equipments_header']['border']

    for width in width_columns:
        sheet.column_dimensions[width].width = width_columns[width]
